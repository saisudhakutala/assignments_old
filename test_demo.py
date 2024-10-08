import pytest
from fastapi.testclient import TestClient
import os
from openpyxl import Workbook
from demo import app

def temp_excel_file():
        file_name = "test_excel.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["name", "age", "email"])  # Adding headers
        ws.append(["John", 25, "john@example.com"])
        ws.append(["Alice", 30, "alice@example.com"])
        wb.save(file_name)

# Initialize the test client
client = TestClient(app)

# Test GET: /get_data
def test_get_data(temp_excel_file):
    app.excel_file = temp_excel_file

    response = client.get("/get_data")
    assert response.status_code == 200
    data = response.json()

# Test POST: /insert_data
def test_insert_data(temp_excel_file):
    app.excel_file = temp_excel_file

    new_data = {"name": "Bob", "age": 40, "email": "bob@example.com"}
    response = client.post("/insert_data", json=new_data)
    assert response.status_code == 200
    assert response.json() == {"message": "Data inserted successfully."}

# Test PUT: /update_data
def test_update_data(temp_excel_file):
    app.excel_file = temp_excel_file

    update_data = {"row_index": 2, "column_name": "age", "new_value": "35"}
    response = client.put("/update_data", json=update_data)
    assert response.status_code == 200
    wb = Workbook()
    wb.load(temp_excel_file)
    ws = wb.active
    updated_age = ws.cell(row=2 + 1, column=2).value  # Row 2, column "age"
    assert updated_age == 35

# Test DELETE: /delete_data
def test_delete_data(temp_excel_file):
    # Set the temporary file path as the excel file path in the app
    app.excel_file = temp_excel_file

    response = client.delete("/delete_data?key=some_key")
    assert response.status_code == 200
    assert response.json() == {"message": "Excel data cleared."}

    # Verify the data was cleared
    wb = Workbook()
    wb.load(temp_excel_file)
    ws = wb.active
    assert ws.max_row == 1

# Test PATCH: /rename_data
def test_rename_data(temp_excel_file):
    # Set the temporary file path as the excel file path in the app
    app.excel_file = temp_excel_file

    update_column_data = {"row_index": 1, "column_name": "age", "new_value": "32"}
    response = client.patch("/rename_data", json=update_column_data)
    assert response.status_code == 200
    assert response.json() == {"message": "Column 'age' renamed to 'age'."}

    wb = Workbook()
    wb.load(temp_excel_file)
    ws = wb.active
    updated_age = ws.cell(row=2 + 1, column=2).value
    assert updated_age == 32

# run it with pytest test_demo.py
