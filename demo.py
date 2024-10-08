# Develop a REST API and perform Put,Patch,GET,Post and Delete operations to update values in db.
# Assume excel as db table.

# Put -> Create new column
# Patch -> Rename specific columns
# GET -> fetch the excel values in a json
# Post -> Update values in row/column
# Delete -> Empty excel



# An element of array is considered a leader if it is greater than all the elements on its right side or if it is equal to the maximum element
#    on its right side. the rightmost element is always a leader.
#    Input: n = 6, arr = {16,17,4,3,5,2}
#    Output: 17 5 2
 
#    Input: n = 5, arr = {10,4,2,4,1}
#    Output: 10 4 4 1
 
#    Time complexity: O(n)
#    Space complexity: O(n)







from fastapi import FastAPI, HTTPException
from pydantic import BaseModel
from openpyxl import load_workbook
import os

app = FastAPI()

excel_file = "demo.xlsx"

if not os.path.exists(excel_file):
    wb = load_workbook(excel_file)
    ws = wb.active
    ws.append(["name", "age", "email"])  # Adding headers
    wb.save(excel_file)


def load_excel():
    return load_workbook(excel_file)


def save_excel(wb):
    wb.save(excel_file)

# model to add new row or updating existing row
class DataModel(BaseModel):
    name: str
    age: int
    email: str

# model to update the values in a specific row/column
class UpdateModel(BaseModel):
    row_index: int
    column_name: str
    new_value: str

# model for updating specific data in the column
class UpdateColumnDataModel(BaseModel):
    row_index: int
    column_name: str
    new_value: str

@app.get("/get_data")
def get_data():
    wb = load_excel()
    ws = wb.active

    # Extract data from rows and columns
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]  # First row is the header
    data = [dict(zip(headers, row)) for row in rows[1:]]  # Map data to headers

    return data


@app.post("/insert_data")
def insert_data(new_data: DataModel):
    wb = load_excel()
    ws = wb.active
    
    # Insert the new row with the provided data
    ws.append([new_data.name, new_data.age, new_data.email])
    save_excel(wb)
    
    return {"message": "Data inserted successfully."}

@app.put("/update_data")
def update_data(update: UpdateModel):
    wb = load_excel()
    ws = wb.active

    # Get the column headers to ensure the column exists
    headers = [cell.value for cell in ws[1]]

    if update.column_name not in headers:
        raise HTTPException(status_code=404, detail="Column not found.")

    # Find the column index
    col_index = headers.index(update.column_name) + 1  # Openpyxl is 1-based index

    # Update the value in the specified row and column
    if update.row_index < 1 or update.row_index > ws.max_row:
        raise HTTPException(status_code=404, detail="Row index out of bounds.")

    ws.cell(row=update.row_index + 1, column=col_index, value=update.new_value)
    save_excel(wb)

    return {"message": f"Value updated successfully at row {update.row_index}, column '{update.column_name}'."}


@app.delete("/delete_data")
def delete_data(key):
    wb = load_excel()
    ws = wb.active
    
    # Clear all data except the header (row 1)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    save_excel(wb)
    
    return {"message": "Excel data cleared."}


@app.patch("/rename_data")
def rename_data(update_data: UpdateColumnDataModel):
    df = load_excel()
    # Check if column exists
    if update_data.column_name not in df.columns:
        raise HTTPException(status_code=404, detail="Column not found.")
    
    # Check if row index is valid
    if update_data.row_index >= len(df):
        raise HTTPException(status_code=404, detail="Row index out of bounds.")
    
    # Update the value at the specified row and column
    df.at[update_data.row_index, update_data.column_name] = update_data.new_value
    save_excel(df)



if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app)
